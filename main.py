import csv
from string import ascii_uppercase
import argparse, textwrap
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.filters import FilterColumn

import utility

# Note: This looks helpful.
# Openpyxl - How to apply autofilter to columns
# https://openpyxl.readthedocs.io/en/latest/filters.html
def configure_filters(worksheet, data_col_height, cols_amount):
    """ configure filters for the spreadsheet """
    ## NOTE: worksheet -> the worksheet we're working on
    ## NOTE: data_col_height -> number of rows (height) of the particular column
    ## NOTE: cols_amount -> Amount of columns on the worksheet

    filters = worksheet.auto_filter
    # filters.ref = f"A1:C{data_col_height}"
    filters.ref = f"A1:{ascii_uppercase[cols_amount - 1]}{data_col_height}"
    col = FilterColumn(colId = 0)
    filters.filterColumn.append(col)

def set_zoom_scale(workbook):
    for ws in workbook.worksheets:
        ws.sheet_view.zoomScale = 110

def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def bold_header(ws):
    """ Bold every cell in the header row. (Iterating over uppercase letters.) """
    for c in ascii_uppercase:
        ws[f'{c}1'].font = Font(bold=True)


def build_dict(row, filter_cols) -> dict[str, str]:
    """Return a new dictionary of items from the row, that are in the filter columns."""
    # NOTE: https://www.guru99.com/python-dictionary-append.html

    data = {}
    
    for key in row.keys():
        if key in filter_cols:
            data[key] = row[key]

    return data

# This was helpful: https://stackoverflow.com/questions/37182528/how-to-append-data-using-openpyxl-python-to-excel-file-from-a-specified-row
def read_csv(filename: str, filter_cols: list[str]) -> dict[str, dict[str, str]]:
    """ Read CSV into two dictionaries. A full and abridged version. """
    full = []
    abridged = []

    with open(filename, 'r', encoding = 'utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            full.append(row)
            abridged.append(build_dict(row, filter_cols))            
    
    return { 'full': full, 'abridged': abridged }

def create_full_workbook(workbook, worksheet, full: dict[str, str], opt_bold_header: bool, opt_auto_filter: bool, opt_auto_width_columns: bool):
    """Manage creating the 'full' workbook. This one is basically a copy of the original CSV data."""
    workbook.active = worksheet
    num_of_cols = len(full[0].keys())
    col_names = list(full[0].keys())

    worksheet.append(col_names) # NOTE: Creates header row
    if opt_bold_header:
        bold_header(worksheet) # NOTE: Must execute AFTER data has been written to the header row.
    
    full_sheet_row_counter = 0
    for f_row in full:
        worksheet.append( utility.get_row(f_row, col_names) )
        full_sheet_row_counter += 1
    
    if opt_auto_width_columns:
        autofit_columns(worksheet)

    if opt_auto_filter:
        configure_filters(worksheet, full_sheet_row_counter, num_of_cols)

def create_filtered_workbook(workbook, worksheet, abridged: dict[str, str], filter_columns: list[str], opt_bold_header: bool, opt_auto_filter: bool, opt_auto_width_columns: bool):
    workbook.active = worksheet
    worksheet.append(filter_columns)

    if opt_bold_header:
        bold_header(worksheet) # NOTE: Must execute AFTER data has been written to the header row.

    row_counter = 0
    abr_num_of_cols = len(abridged[0].keys())
    abr_col_names = list(abridged[0].keys())
    for row in abridged:
        # TODO: Add ability to *add* columns instead of only filter what is already there.
        # ws1.append([row['Posting Date'], row['Amount'], row['Description'], row['Transaction Category'], row['Extended Description']])
        worksheet.append( utility.get_row_filtered(row, abr_col_names, filter_columns) )
        row_counter += 1
    
    if opt_auto_width_columns:
        autofit_columns(worksheet)
    
    if opt_auto_filter:
        configure_filters(worksheet, row_counter, abr_num_of_cols)

def create_workbooks(
        full: dict[str, str], 
        abridged: dict[str, str], 
        filter_columns: str, 
        sheet_name: str, 
        output: str,
        bold_option,
        auto_filter_option,
        auto_width_option
    ):
    """Create workbooks and control what changes are made to them. (This is where the magic happens)"""

    filter_columns = utility.format_filter_cols(filter_columns)

    wb = Workbook()
    del wb['Sheet'] # Delete the default sheet

    ws1 = wb.create_sheet(sheet_name)
    ws2 = wb.create_sheet('raw_data')
    
    create_filtered_workbook(
        workbook=wb, worksheet=ws1, abridged=abridged, filter_columns=filter_columns, 
        opt_bold_header=bold_option, opt_auto_filter=auto_filter_option, opt_auto_width_columns=auto_width_option)
    create_full_workbook(
        workbook=wb, worksheet=ws2, full=full, 
        opt_bold_header=bold_option, opt_auto_filter=auto_filter_option, opt_auto_width_columns=auto_width_option)

    set_zoom_scale(wb)
    wb.save(output)

def main():
    parser = argparse.ArgumentParser(prog='excelify',
                                     formatter_class=argparse.RawDescriptionHelpFormatter,
                                     epilog=textwrap.dedent('''
                                                            Examples:
                                                                excelify.py --csv [file]
                                                                excelify.py --csv [file] --options bold_header,filter,af_cols
                                                        '''))
    parser.add_argument('--csv', type=str, required=True, dest='arg_csv', help="Path to .csv file to process")
    parser.add_argument('--output', type=str, required=True, dest='arg_output', help="Output path for resulting .xlsx file")
    parser.add_argument('--sheet', type=str, required=True, dest='arg_sheet', help="Worksheet name where filtered data will land")
    parser.add_argument('--filter-cols', type=str, required=True, dest='arg_filter_cols', help="Comma-separated list of columns to INCLUDE on new worksheet, other columns are left behind on 'raw' worksheet")

    parser.add_argument('--no-bh', action=argparse.BooleanOptionalAction, dest='arg_no_bold', help="Turn off bold headers")
    parser.add_argument('--no-af', action=argparse.BooleanOptionalAction, dest='arg_no_auto_filter', help="Turn off auto-filters")
    parser.add_argument('--no-aw', action=argparse.BooleanOptionalAction, dest='arg_no_auto_width', help="Turn off auto-width fitment")
    
    args = parser.parse_args()
    arg_csv = args.arg_csv
    arg_output = args.arg_output
    arg_sheet = args.arg_sheet
    arg_filter_cols = args.arg_filter_cols
    arg_no_bold = args.arg_no_bold
    arg_no_auto_filter = args.arg_no_auto_filter
    arg_no_auto_width = args.arg_no_auto_width

    filter_columns = utility.format_filter_cols(arg_filter_cols)
    datasheets = read_csv(arg_csv, filter_columns)
    create_workbooks(
        full=datasheets['full'], 
        abridged=datasheets['abridged'], 
        filter_columns=arg_filter_cols, 
        sheet_name=arg_sheet, 
        output=arg_output,
        bold_option=utility.parse_optional_bool_flag(arg_no_bold),
        auto_filter_option=utility.parse_optional_bool_flag(arg_no_auto_filter),
        auto_width_option=utility.parse_optional_bool_flag(arg_no_auto_width)
    )

if __name__ == "__main__":
    main()